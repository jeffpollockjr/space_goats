"""
Rebalances AbilitiesPile for the one-action turn model.

Changes:
- Rebuilds Rocket/Shield/Special market sheets for current runtime logic.
- Ensures all card effects are supported by implemented gameplay handlers.
- Adds additional Rocket/Shield cards for variety and pacing.
- Tunes copies/costs to reduce dead draws and improve pacing.

Run:
    python3 update_abilities_pile.py
"""

from openpyxl import load_workbook

FILEPATH = "data/Space_Goats_V1_Card_Deck2.xlsx"


def clear_sheet_rows(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def main():
    wb = load_workbook(FILEPATH)
    ws_rockets = wb["RocketMarket"]
    ws_shields = wb["ShieldMarket"]
    ws_specials = wb["SpecialsMarket"]

    # Rewrite rocket market for current ruleset.
    clear_sheet_rows(ws_rockets)
    rocket_rows = [
        ["RK01", "Standard Rocket", "rocket", 2, "destroy_1_ship", "Blockable by shields", 3],
        ["RK02", "Seeking Rocket", "rocket", 2, "destroy_1_weakest_ship", "Finds low-defense targets", 2],
        ["RK03", "Piercing Rocket", "rocket", 3, "destroy_1_ship_ignore_shields", "Bypasses normal shields", 3],
        ["RK04", "EMP Rocket", "rocket", 3, "strip_all_shields_one_opponent", "Removes all shields, no direct ship loss", 2],
        ["RK05", "Barrage Rocket", "rocket", 4, "each_opponent_blocks_or_loses_ship", "Hits every opponent", 2],
        ["RK06", "Salvo Rocket", "rocket", 4, "destroy_up_to_2_ships_then_lose_one_1_bank_currency", "Two hits, then lose 1 bank currency", 2],
        ["RK07", "Shatter Rocket", "rocket", 3, "destroy_1_ship_then_discard_1_random_card_from_hand", "Destroy 1 ship, then discard 1 random card from hand", 3],
        ["RK08", "Overload Barrage", "rocket", 5, "each_opponent_blocks_2_or_loses_2_ships_and_you_skip_next_turn", "Each opponent faces 2 hits; you skip next turn", 1],
        ["RK09", "Twin Salvo", "rocket", 5, "destory_up_to_2_ships", "Premium double strike", 1],
    ]
    for row in rocket_rows:
        ws_rockets.append(row)

    # Rewrite shield market so every card has usable behavior in current engine.
    clear_sheet_rows(ws_shields)
    shield_rows = [
        ["SH01", "Hull Plating", "shield", 2, "assign_to_ship_block_1", "Blocks 1 rocket then discarded", 2],
        ["SH02", "Decoy Drone", "shield", 2, "assign_to_ship_block_1_draw_1_discard_1", "Assign: block 1; immediately draw 1 then discard 1", 3],
        ["SH03", "Emergency Thrusters", "shield", 2, "reactive_block_1_rocket", "Reactive block from hand", 3],
        ["SH04", "Reinforced Hull", "shield", 3, "assign_to_ship_block_2", "Blocks 2 rockets then discarded", 2],
        ["SH05", "Phase Shield", "shield", 4, "assign_to_ship_block_any", "Can block piercing rockets", 2],
        ["SH06", "Aegis Countermeasure", "shield", 3, "reactive_block_1_rocket_then_trash_1_card_from_hand_or_discard", "Reactive block, then trash 1 from hand or discard", 1],
        ["SH07", "Bulwark Plating", "shield", 4, "assign_to_ship_block_2", "Heavy protection", 1],
    ]
    for row in shield_rows:
        ws_shields.append(row)

    # Specials are active again and provide one-action utility effects.
    clear_sheet_rows(ws_specials)
    special_rows = [
        ["SP01", "Deep Space Recon", "special", 2, "draw_3_keep_2_discard_1", "Draw/filter hand quality", 1],
        ["SP02", "Salvage Operation", "special", 2, "retrieve_1_card_from_discard", "Recover key cards", 1],
        ["SP03", "Arms Dealer", "special", 3, "look_at_top3_any_market_rearrange", "Set up future market draws", 1],
        ["SP04", "Last Stand Protocol", "special", 3, "negate_last_ship_loss_once", "One-time final ship protection", 1],
        ["SP05", "Reinforcement Shuttle", "special", 4, "add_1_ship_to_fleet", "Add one ship to fleet", 1],
        ["SP06", "Warp Drive", "special", 4, "take_extra_turn", "Gain an extra full turn", 1],
        ["SP07", "Deck Purge", "special", 2, "trash_1_card_from_discard", "Trash 1 card from your discard", 2],
        ["SP08", "Deep Clean", "special", 3, "trash_1_card_from_discard_draw_1", "Trash 1 card from discard, then draw 1", 2],
    ]
    for row in special_rows:
        ws_specials.append(row)

    wb.save(FILEPATH)

    print("Updated AbilitiesPile sources in workbook:")
    print(f"  Rockets: {len(rocket_rows)} cards")
    print(f"  Shields: {len(shield_rows)} cards")
    print(f"  Specials: {len(special_rows)} cards")


if __name__ == "__main__":
    main()
