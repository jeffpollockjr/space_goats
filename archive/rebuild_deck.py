"""
Rebuilds Space_Goats_V1_Card_Deck2.xlsx from scratch with the new deck-builder structure.
    python3 rebuild_deck.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Helper ────────────────────────────────────────────────────────────────────

def style_header(ws, row=1, fill_hex="1F4E79"):
    fill = PatternFill("solid", start_color=fill_hex)
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

def set_col_widths(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

def add_rows(ws, rows):
    for row in rows:
        ws.append(row)

# ── 1. StarterDeck ────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "StarterDeck"
ws1.append(["card_id", "name", "type", "effect", "notes", "quantity"])
style_header(ws1, fill_hex="1F4E79")
add_rows(ws1, [
    ["ST01", "Fuel Pod",    "currency", "gain_1_currency",          "Basic economy card",              6],
    ["ST02", "Scout Ship",  "currency", "gain_1_currency_draw_1",   "1 currency + draw a card",        2],
    ["ST03", "Weak Rocket", "rocket",   "destroy_1_unshielded_ship","Cannot pierce shields",           2],
])
set_col_widths(ws1, [8, 18, 10, 30, 35, 10])

# ── 2. RocketMarket ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("RocketMarket")
ws2.append(["card_id", "name", "type", "cost", "effect", "notes", "copies"])
style_header(ws2, fill_hex="C00000")
add_rows(ws2, [
    ["RK01", "Standard Rocket", "rocket", 2, "destroy_1_ship",                    "Blockable by any shield",              3],
    ["RK02", "Seeking Rocket",  "rocket", 2, "destroy_1_weakest_ship",             "Targets opponent's least-shielded ship", 3],
    ["RK03", "Piercing Rocket", "rocket", 3, "destroy_1_ship_ignore_shields",      "Ignores all shields",                  3],
    ["RK04", "EMP Rocket",      "rocket", 3, "strip_all_shields_one_opponent",     "Removes shields, no ship loss",        3],
    ["RK05", "Barrage Rocket",  "rocket", 4, "each_opponent_blocks_or_loses_ship", "Hits every opponent",                  3],
    ["RK06", "Salvo Rocket",    "rocket", 4, "destroy_up_to_2_ships",              "Fire twice at one or two targets",     3],
])
set_col_widths(ws2, [8, 18, 10, 6, 35, 40, 8])

# ── 3. ShieldMarket ───────────────────────────────────────────────────────────
ws3 = wb.create_sheet("ShieldMarket")
ws3.append(["card_id", "name", "type", "cost", "effect", "notes", "copies"])
style_header(ws3, fill_hex="375623")
add_rows(ws3, [
    ["SH01", "Hull Plating",         "shield", 2, "assign_to_ship_block_1",         "Blocks 1 rocket then permanently discarded",   3],
    ["SH02", "Decoy Drone",          "shield", 2, "redirect_rocket_to_own_ship",    "Reactive: redirect rocket to a ship you choose", 3],
    ["SH03", "Emergency Thrusters",  "shield", 2, "cancel_1_rocket_targeting_you",  "Reactive: cancel one incoming rocket this turn", 3],
    ["SH04", "Reinforced Hull",      "shield", 3, "assign_to_ship_block_2",         "Blocks 2 rockets then permanently discarded",  3],
    ["SH05", "Phase Shield",         "shield", 4, "assign_to_ship_block_any",       "Blocks all rocket types including Piercing",   3],
])
set_col_widths(ws3, [8, 22, 10, 6, 35, 50, 8])

# ── 4. SpecialsMarket ─────────────────────────────────────────────────────────
ws4 = wb.create_sheet("SpecialsMarket")
ws4.append(["card_id", "name", "type", "cost", "effect", "notes", "copies"])
style_header(ws4, fill_hex="7030A0")
add_rows(ws4, [
    ["SP01", "Deep Space Recon",      "special", 2, "draw_3_keep_2_discard_1",           "Card selection / deck thinning",        2],
    ["SP02", "Salvage Operation",     "special", 2, "retrieve_1_card_from_discard",       "Pull any card from your discard pile",  2],
    ["SP03", "Arms Dealer",           "special", 3, "look_at_top3_any_market_rearrange",  "Manipulate the market row order",       2],
    ["SP04", "Last Stand Protocol",   "special", 3, "negate_last_ship_loss_once",         "One-time save when down to 1 ship",     2],
    ["SP05", "Reinforcement Shuttle", "special", 4, "add_1_ship_to_fleet",                "Gain a new ship immediately",           2],
    ["SP06", "Warp Drive",            "special", 4, "take_extra_turn",                    "Play an additional full turn now",      2],
])
set_col_widths(ws4, [8, 22, 10, 6, 38, 45, 8])

# ── 5. Rules ──────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Rules")
ws5.append(["section", "rule", "value"])
style_header(ws5, fill_hex="2E4057")
add_rows(ws5, [
    ["Setup",    "Starting ships (2 players)",  "5"],
    ["Setup",    "Starting ships (3 players)",  "4"],
    ["Setup",    "Starting ships (4+ players)", "3"],
    ["Setup",    "Starter deck size",           "10 cards (identical for all players)"],
    ["Turn",     "Step 1 - Draw",               "Draw until hand has 5 cards"],
    ["Turn",     "Step 2 - Play",               "Play any cards from hand in any order"],
    ["Turn",     "Step 3 - Buy",                "Spend currency to buy from market rows (top card of each row visible)"],
    ["Turn",     "Step 4 - Cleanup",            "Discard remaining hand to personal discard pile"],
    ["Currency", "How it works",                "Currency cards played this turn are spent — cannot be saved to next turn"],
    ["Currency", "Bought cards go to",          "Your personal discard pile (not hand)"],
    ["Shields",  "When destroyed",              "Permanently discarded — do NOT return to player's deck"],
    ["Market",   "Buying",                      "Only the top card of each market row is available to buy"],
    ["Market",   "Rocket row copies",           "3 copies of each card"],
    ["Market",   "Shield row copies",           "3 copies of each card"],
    ["Market",   "Specials row copies",         "2 copies of each card"],
    ["Hand",     "Maximum hand size",           "5 cards"],
])
set_col_widths(ws5, [12, 35, 50])

filepath = "data/Space_Goats_V1_Card_Deck2.xlsx"
wb.save(filepath)
print(f"Saved to {filepath}")
print()
print("Sheets created:")
for ws in wb.worksheets:
    print(f"  {ws.title}")
print()
print("Starter deck:")
for row in wb["StarterDeck"].iter_rows(min_row=2, values_only=True):
    if row[0]: print(f"  {row[1]:20s} x{row[5]}  —  {row[3]}")
print()
print("Rocket market:")
for row in wb["RocketMarket"].iter_rows(min_row=2, values_only=True):
    if row[0]: print(f"  {row[1]:22s} cost:{row[3]}  x{row[6]}  —  {row[4]}")
print()
print("Shield market:")
for row in wb["ShieldMarket"].iter_rows(min_row=2, values_only=True):
    if row[0]: print(f"  {row[1]:22s} cost:{row[3]}  x{row[6]}  —  {row[4]}")
print()
print("Specials market:")
for row in wb["SpecialsMarket"].iter_rows(min_row=2, values_only=True):
    if row[0]: print(f"  {row[1]:22s} cost:{row[3]}  x{row[6]}  —  {row[4]}")
