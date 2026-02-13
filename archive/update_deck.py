"""
Run this once to update your Excel deck file with the new card quantities
and set all purchase card costs to 3.
    python3 update_deck.py
"""
from openpyxl import load_workbook
import os

filepath = "data/Space_Goats_V1_Card_Deck2.xlsx"

if not os.path.exists(filepath):
    print(f"ERROR: Could not find {filepath}")
    print("Make sure you're running this from the space_goats_sim folder.")
    exit(1)

wb = load_workbook(filepath)

# ── DrawPile: revised quantities ──────────────────────────────────────────────
draw_changes = {
    'R01': 5,   # Fuel Cell           was 12
    'R02': 4,   # Salvage Crate       was 10
    'R03': 4,   # Smuggled Parts      was 8
    'R04': 2,   # Emergency Supplies  was 6
    'S01': 4,   # Scrap Shield        was 8
    'S02': 3,   # Emergency Shield    was 6
    'S03': 2,   # Decoy Drone         was 4
    'K01': 8,   # Basic Rocket        was 4
    'K02': 6,   # Pressure Rocket     was 2
}

ws_draw = wb['DrawPile']
for row in ws_draw.iter_rows(min_row=2):
    card_id = row[1].value
    if card_id in draw_changes:
        old = row[6].value
        row[6].value = draw_changes[card_id]
        print(f"  DrawPile  {card_id:4s} {row[2].value:25s} qty: {old} → {draw_changes[card_id]}")

# ── PurchasePile: all costs → 3 ───────────────────────────────────────────────
ws_purchase = wb['PurchasePile']
for row in ws_purchase.iter_rows(min_row=2):
    if row[0].value is not None:
        old = row[4].value
        row[4].value = 3
        print(f"  Purchase  {row[1].value:4s} {row[2].value:25s} cost: {old} → 3")

wb.save(filepath)
print(f"\nDone! Saved to {filepath}")

# Quick summary
wb2 = load_workbook(filepath)
draw_total = sum(row[6].value for row in wb2['DrawPile'].iter_rows(min_row=2) if row[6].value)
currency = sum(row[6].value for row in wb2['DrawPile'].iter_rows(min_row=2)
               if row[3].value == 'resource' and row[6].value)
rockets  = sum(row[6].value for row in wb2['DrawPile'].iter_rows(min_row=2)
               if row[3].value == 'rocket' and row[6].value)
shields  = sum(row[6].value for row in wb2['DrawPile'].iter_rows(min_row=2)
               if row[3].value == 'shield' and row[6].value)
print(f"\nNew draw deck: {draw_total} cards total")
print(f"  Currency: {currency} ({100*currency//draw_total}%)")
print(f"  Rockets:  {rockets}  ({100*rockets//draw_total}%)")
print(f"  Shields:  {shields}  ({100*shields//draw_total}%)")
