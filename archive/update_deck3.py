"""
Updates the starter deck to add Emergency Shield and Debris Field,
and reduces Fuel Pod to 5 and Scout Ship to 1.
    python3 update_deck3.py
"""
from openpyxl import load_workbook

filepath = "data/Space_Goats_V1_Card_Deck2.xlsx"
wb = load_workbook(filepath)
ws = wb["StarterDeck"]

# Update existing quantities
qty_changes = {"ST01": 5, "ST02": 1}  # Fuel Pod: 6→5, Scout Ship: 2→1
for row in ws.iter_rows(min_row=2):
    if row[0].value in qty_changes:
        old = row[5].value
        row[5].value = qty_changes[row[0].value]
        print(f"  {row[1].value:20s} qty: {old} → {qty_changes[row[0].value]}")

# Add two new cards
new_cards = [
    ["ST04", "Emergency Shield", "shield",   "reactive_block_1_rocket",
     "Play out of turn when attacked — blocks 1 rocket, then discarded", 1],
    ["ST05", "Debris Field",     "debris",   "no_effect",
     "Dead card — clogs hand early game, prune via Salvage/Recon",       1],
]
for card in new_cards:
    ws.append(card)
    print(f"  Added: {card[1]:20s} x{card[5]}")

wb.save(filepath)
print(f"\nDone! New starter deck:")
wb2 = load_workbook(filepath)
total = 0
for row in wb2["StarterDeck"].iter_rows(min_row=2, values_only=True):
    if row[0]:
        print(f"  {row[1]:22s} x{row[5]}  —  {row[3]}")
        total += row[5]
print(f"  Total: {total} cards")
