"""
Run once to remove Phase Barrier and Black Market Dealer from the purchase pile.
    python3 update_deck2.py
"""
from openpyxl import load_workbook

filepath = "data/Space_Goats_V1_Card_Deck2.xlsx"
wb = load_workbook(filepath)
ws = wb['PurchasePile']

remove_ids = {'PS03', 'PE03'}  # Phase Barrier, Black Market Dealer
rows_to_delete = []

for row in ws.iter_rows(min_row=2):
    if row[1].value in remove_ids:
        rows_to_delete.append(row[0].row)
        print(f"  Removing: {row[2].value} ({row[1].value})")

for row_num in reversed(rows_to_delete):
    ws.delete_rows(row_num)

wb.save(filepath)
print(f"\nDone! Remaining purchase cards:")
wb2 = load_workbook(filepath)
for row in wb2['PurchasePile'].iter_rows(min_row=2, values_only=True):
    if row[0]:
        print(f"  {row[1]:4s}  {row[2]:30s}  cost:{row[4]}  qty:{row[7]}")
